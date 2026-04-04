import time, json, os, sys, re
from anthropic import Anthropic
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = "vehicles_nhtsa_1999_2026.xlsx"
PROGRESS_FILE = "refill_progress.json"
MODEL = "claude-haiku-4-5-20251001"
DELAY = 0.15
START_ROW = 26912
SYSTEMS_ENG = ["Engine","Cooling","Electrical","Fuel","Exhaust","Transmission","Suspension","Steering","Brakes","Electronics"]
PROMPT = """For the vehicle: {make} {model} {year}
List the most common known problems for each system below.
Be specific to this exact make, model and year.
Answer in Hebrew. Keep each issue brief (under 10 words).
If no known issues for a system, write "אין בעיות ידועות".
Return ONLY valid JSON, no other text:
{{"Engine":["בעיה1","בעיה2"],"Cooling":["בעיה1"],"Electrical":["בעיה1"],"Fuel":["בעיה1"],"Exhaust":["בעיה1"],"Transmission":["בעיה1"],"Suspension":["בעיה1"],"Steering":["בעיה1"],"Brakes":["בעיה1"],"Electronics":["בעיה1"]}}"""
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', str(text))
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f: return json.load(f)
    return {"last_row": START_ROW - 1}
def save_progress(row):
    with open(PROGRESS_FILE, "w") as f: json.dump({"last_row": row}, f)
def get_issues(client, make, model, year, retries=2):
    for attempt in range(retries + 1):
        try:
            r = client.messages.create(model=MODEL, max_tokens=600,
                messages=[{"role":"user","content":PROMPT.format(make=make,model=model,year=year)}])
            text = r.content[0].text.strip().replace("```json","").replace("```","").strip()
            return json.loads(text)
        except json.JSONDecodeError:
            if attempt < retries: time.sleep(1); continue
            return None
        except Exception as e:
            if "rate" in str(e).lower() and attempt < retries:
                print(f"    Rate limited, waiting 30s..."); time.sleep(30); continue
            print(f"    API error: {e}"); return None
def row_needs_refill(ws, row):
    """Check if row has empty or 'no data' content"""
    for col in range(4, 14):
        val = ws.cell(row=row, column=col).value
        if val and str(val).strip() not in ["", "אין נתונים", "אין בעיות ידועות"]:
            return False
    return True
def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        api_key = input("Enter your Anthropic API key (sk-ant-...): ").strip()
    if not api_key: print("No API key. Exiting."); sys.exit(1)
    client = Anthropic(api_key=api_key)
    print("Loading Excel...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    total_rows = ws.max_row
    print(f"Total rows: {total_rows}")
    print(f"Scanning rows {START_ROW} to {total_rows} for empty/no-data rows...")
    progress = load_progress()
    start_row = progress["last_row"] + 1
    if start_row > START_ROW: print(f"Resuming from row {start_row}")
    total_range = total_rows - start_row + 1
    errors = 0; processed = 0; skipped = 0
    for row in range(start_row, total_rows + 1):
        make = ws.cell(row=row, column=1).value
        model = ws.cell(row=row, column=2).value
        year = ws.cell(row=row, column=3).value
        if not make or not model or not year: continue
        if not row_needs_refill(ws, row):
            skipped += 1
            continue
        make_s, model_s, year_i = str(make).strip(), str(model).strip(), int(year)
        issues = get_issues(client, make_s, model_s, year_i)
        df = Font(name="Arial",size=9)
        wa = Alignment(wrap_text=True,vertical="top",horizontal="right")
        br = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
        alt = PatternFill("solid",fgColor="F2F7FB") if row%2==0 else PatternFill()
        if issues:
            for j, sys_eng in enumerate(SYSTEMS_ENG):
                col = 4 + j
                data = issues.get(sys_eng, ["אין נתונים"])
                text = " | ".join(clean_text(x) for x in data) if isinstance(data, list) else clean_text(str(data))
                c = ws.cell(row=row, column=col, value=text)
                c.font=df; c.alignment=wa; c.border=br; c.fill=alt
            processed += 1
        else:
            errors += 1
        done = row - START_ROW + 1
        if done%50==0 or row==total_rows:
            pct=(done/total_range)*100
            print(f"  [{done}/{total_range}] ({pct:.1f}%) {make_s} {model_s} {year_i} | Filled:{processed} Skip:{skipped} Err:{errors}")
        if done%500==0:
            save_progress(row); wb.save(EXCEL_FILE)
            print(f"  >> Saved at row {row}")
        elif done%100==0:
            save_progress(row)
        time.sleep(DELAY)
    print(f"\n=== Done! Filled:{processed} Skipped:{skipped} Errors:{errors} ===")
    wb.save(EXCEL_FILE)
    if os.path.exists(PROGRESS_FILE): os.remove(PROGRESS_FILE)
    print(f"Updated {EXCEL_FILE}!")
if __name__ == "__main__":
    main()
