import time, json, os, sys, re
from anthropic import Anthropic
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = "vehicles_nhtsa_1999_2026.xlsx"
PROGRESS_FILE = "israel_vehicles_progress.json"
MODEL = "claude-haiku-4-5-20251001"
DELAY = 0.15
SYSTEMS_ENG = ["Engine","Cooling","Electrical","Fuel","Exhaust","Transmission","Suspension","Steering","Brakes","Electronics"]
PROMPT = """For the vehicle: {make} {model} {year}
List the most common known problems for each system below.
Be specific to this exact make, model and year.
Answer in Hebrew. Keep each issue brief (under 10 words).
If no known issues for a system, write "אין בעיות ידועות".
Return ONLY valid JSON, no other text:
{{"Engine":["בעיה1","בעיה2"],"Cooling":["בעיה1"],"Electrical":["בעיה1"],"Fuel":["בעיה1"],"Exhaust":["בעיה1"],"Transmission":["בעיה1"],"Suspension":["בעיה1"],"Steering":["בעיה1"],"Brakes":["בעיה1"],"Electronics":["בעיה1"]}}"""
ISRAEL_VEHICLES = [
    ("AUDI","A1",2010,2026),("AUDI","Q2",2017,2026),
    ("BMW","1 Series (116i/118i/120i)",2004,2026),("BMW","2 Series (218i/220i)",2014,2026),
    ("BMW","3 Series (316i/318i/320i/330i)",2000,2026),("BMW","4 Series (420i/430i)",2014,2026),
    ("BMW","5 Series (520i/525i/530i)",2000,2026),("BMW","7 Series (730i/740i)",2000,2026),
    ("BYD","Atto 3",2022,2026),("BYD","Dolphin",2023,2026),("BYD","Han",2022,2026),
    ("BYD","Seal",2023,2026),("BYD","Tang",2022,2026),
    ("CHERY","Arrizo 6",2021,2026),("CHERY","Tiggo 4",2021,2026),
    ("CHERY","Tiggo 7",2021,2026),("CHERY","Tiggo 8",2021,2026),
    ("CITROEN","C2",2003,2010),("CITROEN","C8",2002,2014),("CITROEN","Jumpy",2007,2026),
    ("CITROEN","Nemo",2008,2018),("CITROEN","Saxo",2000,2004),("CITROEN","Xsara",2000,2006),
    ("CITROEN","Xsara Picasso",2000,2010),
    ("CUPRA","Ateca",2021,2026),("CUPRA","Born",2022,2026),("CUPRA","Formentor",2021,2026),
    ("CUPRA","Leon",2021,2026),
    ("DACIA","Dokker",2013,2021),("DACIA","Lodgy",2012,2022),
    ("FIAT","500",2007,2026),("FIAT","Bravo",2007,2015),("FIAT","Doblo",2001,2023),
    ("FIAT","Linea",2007,2016),("FIAT","Multipla",2000,2010),("FIAT","Panda",2003,2026),
    ("FIAT","Punto",2000,2019),("FIAT","Qubo",2008,2020),("FIAT","Stilo",2001,2008),
    ("FIAT","Tipo",2016,2026),
    ("GEELY","Atlas",2023,2026),("GEELY","Coolray",2022,2026),("GEELY","Geometry C",2022,2026),
    ("GEELY","Monjaro",2023,2026),
    ("HONDA","City",2003,2015),
    ("HYUNDAI","Atos",2000,2008),("HYUNDAI","Getz",2002,2011),("HYUNDAI","Matrix",2001,2010),
    ("HYUNDAI","Terracan",2001,2008),("HYUNDAI","Venue",2020,2026),
    ("HYUNDAI","ix20",2010,2019),("HYUNDAI","ix35",2010,2016),
    ("KIA","Carens",2000,2019),("KIA","Cerato",2004,2026),("KIA","Venga",2010,2019),
    ("MAZDA","Demio",2000,2008),
    ("MERCEDES-BENZ","EQA",2022,2026),("MERCEDES-BENZ","EQC",2020,2026),
    ("MERCEDES-BENZ","Vito",2000,2026),
    ("MG","HS",2020,2026),("MG","MG4",2023,2026),("MG","MG5",2023,2026),
    ("MG","Marvel R",2022,2026),("MG","ZS",2020,2026),
    ("MITSUBISHI","ASX",2010,2026),("MITSUBISHI","Carisma",2000,2004),
    ("MITSUBISHI","Colt",2000,2013),("MITSUBISHI","Grandis",2004,2011),
    ("MITSUBISHI","L200",2005,2026),("MITSUBISHI","Pajero",2000,2021),
    ("MITSUBISHI","Space Star",2013,2026),
    ("NISSAN","Almera",2000,2007),("NISSAN","Navara",2005,2026),("NISSAN","Patrol",2000,2026),
    ("NISSAN","Primera",2000,2008),("NISSAN","Pulsar",2014,2019),("NISSAN","Tiida",2004,2013),
    ("OPEL","Adam",2013,2019),("OPEL","Agila",2000,2015),("OPEL","Astra",2000,2026),
    ("OPEL","Corsa",2000,2026),("OPEL","Crossland",2017,2026),("OPEL","Grandland",2018,2026),
    ("OPEL","Insignia",2009,2023),("OPEL","Karl",2015,2020),("OPEL","Meriva",2003,2017),
    ("OPEL","Mokka",2012,2026),("OPEL","Vectra",2000,2009),("OPEL","Zafira",2000,2019),
    ("PEUGEOT","106",2000,2004),("PEUGEOT","306",2000,2003),("PEUGEOT","307",2001,2009),
    ("PEUGEOT","406",2000,2005),("PEUGEOT","407",2004,2012),("PEUGEOT","607",2000,2012),
    ("PEUGEOT","Boxer",2006,2026),("PEUGEOT","Expert",2007,2026),
    ("RENAULT","Espace",2000,2023),("RENAULT","Laguna",2000,2015),
    ("RENAULT","Latitude",2011,2016),("RENAULT","Modus",2004,2013),
    ("RENAULT","Symbol",2002,2013),("RENAULT","Twingo",2000,2014),
    ("SEAT","Cordoba",2000,2010),
    ("SKODA","Citigo",2012,2021),
    ("SUZUKI","Liana",2001,2008),("SUZUKI","Wagon R",2000,2008),
    ("TOYOTA","Avensis",2000,2019),("TOYOTA","Corolla Cross",2022,2026),
    ("TOYOTA","Hilux",2005,2026),("TOYOTA","Urban Cruiser",2009,2014),
    ("TOYOTA","Verso",2009,2018),
    ("VOLKSWAGEN","Bora",2000,2006),("VOLKSWAGEN","ID.5",2022,2026),
    ("VOLKSWAGEN","Sharan",2000,2023),("VOLKSWAGEN","Taigo",2022,2026),
    ("VOLKSWAGEN","Transporter",2000,2026),
]
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', str(text))
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f: return json.load(f)
    return {"last_index": -1, "added": False}
def save_progress(idx, added=True):
    with open(PROGRESS_FILE, "w") as f: json.dump({"last_index": idx, "added": added}, f)
def get_issues(client, make, model, year, retries=2):
    for attempt in range(retries + 1):
        try:
            r = client.messages.create(model=MODEL, max_tokens=600,
                messages=[{"role":"user","content":PROMPT.format(make=make,model=model,year=year)}])
            text = r.content[0].text.strip().replace("```json","").replace("```","").strip()
            return json.loads(text)
        except json.JSONDecodeError:
            if attempt < retries: time.sleep(1); continue
            return None
        except Exception as e:
            if "rate" in str(e).lower() and attempt < retries:
                print(f"    Rate limited, waiting 30s..."); time.sleep(30); continue
            print(f"    API error: {e}"); return None
def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        api_key = input("Enter your Anthropic API key (sk-ant-...): ").strip()
    if not api_key: print("No API key. Exiting."); sys.exit(1)
    client = Anthropic(api_key=api_key)
    all_rows = []
    for make, model, start, end in ISRAEL_VEHICLES:
        for year in range(start, end + 1):
            all_rows.append((make, model, year))
    print(f"Total Israel vehicles to add: {len(all_rows)}")
    progress = load_progress()
    start_idx = progress["last_index"] + 1
    print("Loading Excel...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    current_last_row = ws.max_row
    if not progress.get("added", False):
        print(f"Adding {len(all_rows)} rows to Excel...")
        df = Font(name="Arial", size=10)
        br = Border(left=Side("thin",color="D9D9D9"),right=Side("thin",color="D9D9D9"),
                    top=Side("thin",color="D9D9D9"),bottom=Side("thin",color="D9D9D9"))
        for i, (make, model, year) in enumerate(all_rows):
            row = current_last_row + 1 + i
            alt = PatternFill("solid",fgColor="F2F7FB") if row%2==0 else PatternFill()
            for col, val in enumerate([make, model, year], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font=df; c.border=br; c.fill=alt
                if col==3: c.alignment=Alignment(horizontal="center"); c.number_format="0"
        wb.save(EXCEL_FILE)
        save_progress(-1, added=True)
        print(f"Added! Now filling with AI...")
        start_idx = 0
    if start_idx > 0: print(f"Resuming from #{start_idx}")
    first_new_row = current_last_row + 1
    errors = 0; processed = 0
    for i in range(start_idx, len(all_rows)):
        make, model, year = all_rows[i]
        row = first_new_row + i
        issues = get_issues(client, make, model, year)
        df = Font(name="Arial",size=9)
        wa = Alignment(wrap_text=True,vertical="top",horizontal="right")
        br = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
        alt = PatternFill("solid",fgColor="F2F7FB") if row%2==0 else PatternFill()
        if issues:
            for j, sys_eng in enumerate(SYSTEMS_ENG):
                col = 4 + j
                data = issues.get(sys_eng, ["אין נתונים"])
                text = " | ".join(clean_text(x) for x in data) if isinstance(data, list) else clean_text(str(data))
                c = ws.cell(row=row, column=col, value=text)
                c.font=df; c.alignment=wa; c.border=br; c.fill=alt
            processed += 1
        else:
            for j in range(10):
                c = ws.cell(row=row, column=4+j, value="אין נתונים")
                c.font=df; c.alignment=wa; c.border=br; c.fill=alt
            errors += 1
        if (i+1)%25==0 or i==len(all_rows)-1:
            pct=((i+1)/len(all_rows))*100
            print(f"  [{i+1}/{len(all_rows)}] ({pct:.1f}%) {make} {model} {year} | OK:{processed} Err:{errors}")
        if (i+1)%200==0:
            save_progress(i,added=True); wb.save(EXCEL_FILE)
            print(f"  >> Saved at #{i+1}")
        time.sleep(DELAY)
    print(f"\n=== Done! Added {processed} vehicles with AI data ===")
    wb.save(EXCEL_FILE)
    if os.path.exists(PROGRESS_FILE): os.remove(PROGRESS_FILE)
    print(f"Updated {EXCEL_FILE}!")
if __name__ == "__main__":
    main()
