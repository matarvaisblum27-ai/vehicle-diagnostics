import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = "vehicles_nhtsa_1999_2026.xlsx"

# Hebrew to English make mapping
MAKE_MAP = {
    "אאודי": "AUDI", "אודי": "AUDI",
    "אופל": "OPEL",
    "אורה": "ORA",
    "איווייס": "AIWAYS",
    "איסוזו": "ISUZU",
    "אלפא רומיאו": "ALFA ROMEO",
    "אסטון מרטין": "ASTON MARTIN",
    "ב מ וו": "BMW",
    "באייק": "BAIC",
    "בי ווי די": "BYD",
    "ביואיק": "BUICK",
    "בנטלי": "BENTLEY",
    "ג'אק": "JAC",
    "ג'אקו": "JACO",
    "ג'יפ": "JEEP",
    "גילי": "GEELY",
    "גרייט וול": "GREAT WALL",
    "דאציה": "DACIA",
    "דודג'": "DODGE",
    "דונגפנג": "DONGFENG",
    "דייהו": "DAEWOO",
    "דייהטסו": "DAIHATSU",
    "הונדה": "HONDA",
    "וולבו": "VOLVO",
    "זיקר": "ZEEKR",
    "טויוטה": "TOYOTA",
    "טסלה": "TESLA",
    "יגואר": "JAGUAR",
    "יונדאי": "HYUNDAI",
    "לוטוס": "LOTUS",
    "לינק אנד קו": "LYNK AND CO",
    "לינקולן": "LINCOLN",
    "למבורגיני": "LAMBORGHINI",
    "לנדרובר": "LAND ROVER",
    "לנצ'יה": "LANCIA", "לנציה": "LANCIA",
    "לקסוס": "LEXUS",
    "מ.ג": "MG",
    "מזדה": "MAZDA",
    "מזארטי": "MASERATI",
    "מיני": "MINI",
    "מיצובישי": "MITSUBISHI",
    "מקלארין": "MCLAREN",
    "מרצדס בנץ": "MERCEDES-BENZ", "מרצדס-בנץ": "MERCEDES-BENZ",
    "ניסאן": "NISSAN",
    "סאאב": "SAAB",
    "סאנגיונג": "SSANGYONG",
    "סובארו": "SUBARU",
    "סוזוקי": "SUZUKI", "סוזוקי-מרוטי": "SUZUKI", "מרוטי-סוזוקי": "SUZUKI",
    "סיאט": "SEAT",
    "סיטרואן": "CITROEN",
    "סמארט": "SMART",
    "סקודה": "SKODA",
    "פולסטאר": "POLESTAR",
    "פולקסווגן": "VOLKSWAGEN",
    "פורד": "FORD",
    "פורשה": "PORSCHE",
    "פורתינג": "FORTHING",
    "פיאט": "FIAT", "פיאט קרייזלר": "FIAT",
    "פיג'ו": "PEUGEOT", "פיגו": "PEUGEOT",
    "פרארי": "FERRARI",
    "צ'רי": "CHERY",
    "קאדילאק": "CADILLAC",
    "קופרה": "CUPRA",
    "קיה": "KIA",
    "קרייזלר": "CHRYSLER",
    "רולס-רויס": "ROLLS-ROYCE",
    "רנו": "RENAULT",
    "שברולט": "CHEVROLET",
    "אומודה": "OMODA",
    "גי.אי.סי": "GAC", "גיאיוואן": "GAC", "גיי.איי.סי": "GAC",
    "האמר": "HUMMER",
    "דיפאל": "DEEPAL",
    "ליפמוטור": "LEAPMOTOR",
    "מקסוס": "MAXUS",
    "נטע": "NETA",
    "ניאו": "NIO",
    "סנטרו": "SUNTOUR",
    "פוטון": "FOTON",
    "אקסלנטיקס": "EXCELLENTICS",
    "אקספנג": "XPENG",
    "ארקפוקס": "ARCFOX",
    "טלקו": "TELCO",
    "לינקסיס": "LINKSYS",
    "ריהיי": "RIHAY",
    "סרס": "SERES",
    "פאריזון": "FARIZON",
    "סקיוול": "SKYWELL",
    "רובר": "ROVER",
    "מורגן": "MORGAN",
    "אלפין": "ALPINE",
    "די.אס": "DS", "די אס": "DS",
    "וואי": "WAY",
    "וויה": "VIA",
    "יודו": "YUDO",
    "קאן": "QUAN",
    "קארמה": "KARMA",
}

def heb_to_eng(heb_make):
    """Convert Hebrew make name to English"""
    heb = heb_make.strip()
    # Remove country suffix (e.g. "טויוטה יפן" -> "טויוטה")
    for key in sorted(MAKE_MAP.keys(), key=len, reverse=True):
        if heb.startswith(key):
            return MAKE_MAP[key]
    return None

def main():
    # Load cached Israel data
    print("Loading Israel vehicles from cache...")
    with open("israel_fetch_progress.json") as f:
        data = json.load(f)
    raw_vehicles = [tuple(v) for v in data["vehicles"]]
    print(f"Raw Israel vehicles: {len(raw_vehicles)}")

    # Convert Hebrew makes to English
    converted = []
    unmapped = set()
    for heb_make, model, year in raw_vehicles:
        eng_make = heb_to_eng(heb_make)
        if eng_make:
            converted.append((eng_make, model, year))
        else:
            unmapped.add(heb_make)

    print(f"Converted to English: {len(converted)}")
    if unmapped:
        print(f"Unmapped makes ({len(unmapped)}):")
        for u in sorted(unmapped): print(f"  {u}")

    # Remove duplicates
    converted = sorted(set(converted), key=lambda x: (x[0], x[1], x[2]))
    print(f"After dedup: {len(converted)}")

    # Load existing Excel
    print("Loading existing Excel...")
    df = pd.read_excel(EXCEL_FILE, usecols=[0,1,2])
    df.columns = ["Make","Model","Year"]
    existing = set()
    for _, row in df.iterrows():
        m = str(row["Make"]).strip().upper()
        d = str(row["Model"]).strip().upper()
        y = int(row["Year"]) if pd.notna(row["Year"]) else 0
        existing.add((m, d, y))
    print(f"Existing in Excel: {len(existing)}")

    # Find missing
    missing = []
    for make, model, year in converted:
        if (make.upper(), model.upper(), year) not in existing:
            missing.append((make, model, year))

    print(f"\nMissing from Excel: {len(missing)}")

    if missing:
        # Show sample by make
        from collections import Counter
        make_counts = Counter(m for m, _, _ in missing)
        print(f"\nMissing by make:")
        for make, count in make_counts.most_common(30):
            print(f"  {make}: {count} vehicles")

        print(f"\nSample missing:")
        for v in missing[:15]: print(f"  {v[0]} {v[1]} {v[2]}")

        answer = input(f"\nAdd {len(missing)} vehicles to Excel? (y/n): ")
        if answer.lower() == 'y':
            print(f"Adding {len(missing)} vehicles...")
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            last_row = ws.max_row
            df_font = Font(name="Arial", size=10)
            br = Border(left=Side("thin",color="D9D9D9"),right=Side("thin",color="D9D9D9"),
                        top=Side("thin",color="D9D9D9"),bottom=Side("thin",color="D9D9D9"))
            for i, (make, model, year) in enumerate(missing):
                row = last_row + 1 + i
                alt = PatternFill("solid",fgColor="F2F7FB") if row%2==0 else PatternFill()
                for col, val in enumerate([make, model, year], 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font=df_font; c.border=br; c.fill=alt
                    if col==3: c.alignment=Alignment(horizontal="center"); c.number_format="0"
            wb.save(EXCEL_FILE)
            print(f"Done! Excel now has {last_row + len(missing) - 1} vehicles.")
        else:
            print("Cancelled.")
    else:
        print("No missing vehicles!")

if __name__ == "__main__":
    main()
