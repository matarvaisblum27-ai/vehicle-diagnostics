import requests, time, json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

API_URL = "https://data.gov.il/api/3/action/datastore_search"
RESOURCE_ID = "053cea08-09bc-40ec-8f7a-156f0677aff3"
EXCEL_FILE = "vehicles_nhtsa_1999_2026.xlsx"
PROGRESS_FILE = "israel_fetch_progress.json"

def load_progress():
    try:
        with open(PROGRESS_FILE) as f: return json.load(f)
    except: return {"vehicles": [], "done_fetch": False}

def save_progress(data):
    with open(PROGRESS_FILE, "w") as f: json.dump(data, f, ensure_ascii=False)

def fetch_unique_vehicles():
    progress = load_progress()
    if progress.get("done_fetch"):
        print(f"Using cached data: {len(progress['vehicles'])} vehicles")
        return [tuple(v) for v in progress["vehicles"]]
    
    vehicles = set()
    offset = 0
    limit = 32000
    print("Fetching from data.gov.il...")
    while True:
        params = {"resource_id": RESOURCE_ID, "limit": limit, "offset": offset,
                  "fields": "tozeret_nm,kinuy_mishari,shnat_yitzur"}
        try:
            resp = requests.get(API_URL, params=params, timeout=60)
            data = resp.json()
            records = data.get("result", {}).get("records", [])
            if not records: break
            for r in records:
                make = str(r.get("tozeret_nm","")).strip().upper()
                model = str(r.get("kinuy_mishari","")).strip()
                year = r.get("shnat_yitzur")
                if make and model and year:
                    try:
                        y = int(year)
                        if 1999 <= y <= 2026: vehicles.add((make, model, y))
                    except: pass
            total = data.get("result",{}).get("total",0)
            offset += limit
            print(f"  Fetched {offset} records... Unique: {len(vehicles)}")
            if offset >= total: break
            time.sleep(1)
        except Exception as e:
            print(f"  Error: {e}"); break
    
    vlist = sorted(vehicles, key=lambda x: (x[0], x[1], x[2]))
    save_progress({"vehicles": vlist, "done_fetch": True})
    return vlist

def find_missing(all_israel):
    print("Loading Excel with pandas (fast)...")
    df = pd.read_excel(EXCEL_FILE, usecols=[0,1,2])
    df.columns = ["Make","Model","Year"]
    existing = set()
    for _, row in df.iterrows():
        m = str(row["Make"]).strip().upper()
        d = str(row["Model"]).strip().upper()
        y = int(row["Year"]) if pd.notna(row["Year"]) else 0
        existing.add((m, d, y))
    print(f"  Existing in Excel: {len(existing)}")
    
    missing = []
    for make, model, year in all_israel:
        if (make.upper(), model.upper(), year) not in existing:
            missing.append((make, model, year))
    return missing

def add_to_excel(missing):
    print(f"Adding {len(missing)} vehicles to Excel...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    last_row = ws.max_row
    df = Font(name="Arial", size=10)
    br = Border(left=Side("thin",color="D9D9D9"),right=Side("thin",color="D9D9D9"),
                top=Side("thin",color="D9D9D9"),bottom=Side("thin",color="D9D9D9"))
    for i, (make, model, year) in enumerate(missing):
        row = last_row + 1 + i
        alt = PatternFill("solid",fgColor="F2F7FB") if row%2==0 else PatternFill()
        for col, val in enumerate([make, model, year], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font=df; c.border=br; c.fill=alt
            if col==3: c.alignment=Alignment(horizontal="center"); c.number_format="0"
    wb.save(EXCEL_FILE)
    print(f"Done! Excel now has {last_row + len(missing) - 1} vehicles.")

def main():
    all_israel = fetch_unique_vehicles()
    print(f"\nTotal unique Israel vehicles (1999-2026): {len(all_israel)}")
    missing = find_missing(all_israel)
    print(f"Missing from Excel: {len(missing)}")
    if missing:
        print(f"\nSample missing:")
        for v in missing[:15]: print(f"  {v[0]} {v[1]} {v[2]}")
        print(f"  ...")
        answer = input(f"\nAdd {len(missing)} vehicles? (y/n): ")
        if answer.lower() == 'y': add_to_excel(missing)
        else: print("Cancelled.")
    else:
        print("No missing vehicles!")

if __name__ == "__main__":
    main()
